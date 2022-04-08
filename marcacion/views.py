from django.views.decorators import gzip
from django.http import StreamingHttpResponse
from django.http.response import HttpResponse
from django.shortcuts import redirect, render
from openpyxl.styles.borders import Side
from openpyxl.styles.fills import PatternFill
from openpyxl.workbook.workbook import Workbook
from empleado.models import*
from horarios.models import *
from datetime import date, datetime,timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment,Border
import os,cv2
from marcacion.monitor import Reconocimientofacial
from marcacion.marcacion import monitor_ausencia,dia, registrarentrada
from marcacion.models import *
# Create your views here.
def menu_marcaciones(request):
    
    return render(request,'marcacion/menu_marcaciones/menu_marcaciones.html')

def listar_marcacion(request):
    w = datetime.now()
    fecha = w.strftime('%Y-%m-%d')
    lmarcacion=Marcacion.objects.filter(fecha_marcacion=fecha)
    return render(request,'marcacion/listar_marcacion/listar_marcacion.html',{'lmarcacion':lmarcacion})

def marcacionmanualentrada(request):
    if request.method=='POST':
        carnet=request.POST['carnet']
        msg=""
        id_empleado=Empleado.objects.filter(carnet_empleado=carnet).first() 
        if id_empleado:  
            if id_empleado.es_activo==True:
                horario=AsigHorario.objects.filter(empleado_id=id_empleado.id)
                now = datetime.now()
                fecha = now.strftime('%Y-%m-%d')
                if horario:
                    for hora in horario:
                        if str(hora.fecha_inicio)<=str(fecha):
                            if str(hora.fecha_fin)>=str(fecha):
                                nueva_entrada=[]
                                turno = Turno.objects.filter(horario_id=hora.horario_id)
                                # si esta dentro del rango consultamos a turno los horarios de acuerdo al dia que se esta consultando
                                # en esta linea obtenemos el dia del dia
                                for t in turno:
                                    # recorremos turno debido a que hay varios horarios con ese 
                                    if t.dia == dia():
                                        # si el turno es igual a dia consultamos 
                                        ho = RangoHora.objects.get(id=t.rango_hora_id)
                                        #consultamos las horas
                                        
                                        tiempo=timedelta(hours=now.hour,minutes=now.minute,seconds=now.second)
                                        hora_add = timedelta(minutes=30)
                                        hora_add_salida = timedelta(minutes=55)                   
                                        ini=ho.hora_inicio
                                        fin=ho.hora_fin
                                        hora_ini=timedelta(hours=ini.hour, minutes=ini.minute, seconds=ini.second)
                                        hora_fin=timedelta(hours=fin.hour, minutes=fin.minute, seconds=fin.second)
                                            
                                        if str(tiempo.seconds)>=str(hora_ini.seconds-hora_add.seconds):                                        
                                            if str(tiempo.seconds)<=str(hora_fin.seconds+hora_add_salida.seconds):
                                                #se agrega una hora a la hora de salida para la vigencia de la marcacion
                                                #si la hora que se consulta esta dentro del rango de la hora a marcar
                                                minutos_fin_mas=timedelta(hours=0, minutes=0, seconds=0)
                                                minutos_fin_menos=timedelta(hours=0, minutes=0, seconds=0)
                                                minutos_ini_mas=timedelta(hours=0, minutes=0, seconds=0)
                                                minutos_ini_menos=timedelta(hours=0, minutes=0, seconds=0)
                                                entrada_tardia=False
                                                salida_temprana=False
                                                # al realizar la resta cuando el resultado suele ser menos hay que hacer una
                                                # con timedelta resta
                                                nueva_entrada.append(ho.rango_hora)
                                                if str(tiempo-hora_ini)< "0":
                                                    minutos_ini_mas=+(timedelta()-(tiempo-hora_ini))  
                                                else:
                                                    minutos_ini_menos=tiempo-hora_ini
                                                    entrada_tardia=True                                                                    
                                                if str(tiempo-hora_fin)<"0":
                                                    minutos_fin_menos=+(timedelta()-(tiempo-hora_fin))
                                                    salida_temprana=True                                                                    
                                                else:
                                                    minutos_fin_mas=tiempo-hora_fin                                                                    
                                                # se consulta la tabla marcacion filtrada por la fecha de la marcacion
                                                # si hay registros del empleado que esta realizando la marcacion esto con el objeto 
                                                # marcar ya sea entrada o salida segun sea el caso
                                                
                                                #realizamos consulta en tabla marcacion para conocer si hay algun registro del dia 
                                                # del empleado
                                                consulta=Marcacion.objects.filter(fecha_marcacion=fecha)
                                                #iteramos la consulta
                                                lista=[]
                                                #esta lista se crea para llenarla con los id de registro de cada empleado
                                                if consulta:
                                                    for iterar in consulta:
                                                        lista.append(str(iterar.empleado_id))                                                    
                                                        if iterar.empleado_id==id_empleado.id:
                                                            if ho.rango_hora==iterar.rango_horario:
                                                                if ho.rango_hora in nueva_entrada:
                                                                    nueva_entrada.remove(ho.rango_hora)
                                                                #si hay un registro con el id del empleado consultamos la hora de entrada
                                                                hora_add_salida = timedelta(minutes=5)#minutos que se le sumaran 
                                                                # a hora entrada para que no sea una marcacion recurrente de entrada por salida
                                                                v_salida=iterar.hora_salida                                                                
                                                                if v_salida:
                                                                    msg="Ya marco la salida de :",iterar.empleado," Rango de hora :",ho.rango_hora
                                                                    #si salida no tiene registro o esta vacio se procedera a realizar la actuyalizacion 
                                                                else:
                                                                    entrada=iterar.hora_entrada
                                                                    hora_entrada=timedelta(hours=entrada.hour, minutes=entrada.minute, seconds=entrada.second)
                                                                    if str(tiempo.seconds)>str(hora_entrada.seconds+hora_add_salida.seconds):
                                                                        #si la hora actual es mayor a la hora de entrada mas los minutos extras definifos
                                                                        #entonces se procedera a realizar la actualizacion de ese registro
                                                                        Marcacion.objects.filter(empleado_id=id_empleado.id).update(
                                                                                                    hora_salida=str(tiempo),
                                                                                                    minutos_salida_mas=minutos_fin_mas,
                                                                                                    minutos_salida_menos=minutos_fin_menos, 
                                                                                                    salida_temprana=salida_temprana )
                                                                        msg="Salida marcada exitosamente de :",iterar.empleado," Rango de hora :",ho.rango_hora
                                                    if not str(id_empleado.id )in lista:
                                                        # si el id no esta en toda lista entonces se crea el registro 
                                                        # de una nueva marcacion
                                                        if ho.rango_hora in nueva_entrada:
                                                            nueva_entrada.remove(ho.rango_hora)
                                                        Marcacion.objects.create(
                                                                        empleado_id = id_empleado.id,
                                                                        
                                                                        horario_id=hora.horario_id,
                                                                        rango_horario=ho.rango_hora,
                                                                        fecha_marcacion=fecha,
                                                                        hora_entrada=str(tiempo),
                                                                        minutos_entrada_mas=str(minutos_ini_mas),
                                                                        minutos_entrada_menos=str(minutos_ini_menos),
                                                                        entrada_tardia=entrada_tardia)
                                                        msg="Entrada creada exitosamente de :",id_empleado," Rango de hora :",ho.rango_hora
                                                    elif ho.rango_hora in nueva_entrada:
                                                        
                                                        Marcacion.objects.create(
                                                                    empleado_id = id_empleado.id,
                                                                    horario_id=hora.horario_id,
                                                                    rango_horario=ho.rango_hora,
                                                                    fecha_marcacion=fecha,
                                                                    hora_entrada=str(tiempo),
                                                                    minutos_entrada_mas=str(minutos_ini_mas),
                                                                    minutos_entrada_menos=str(minutos_ini_menos),
                                                                    entrada_tardia=entrada_tardia)
                                                        msg="Nueva entrada creada exitosamente de :",id_empleado," Rango de hora :",ho.rango_hora
                                                                                
                                                elif not consulta:
                                                    Marcacion.objects.create(
                                                            empleado_id = id_empleado.id,
                                                            horario_id=hora.horario_id,
                                                            rango_horario=str(ho.rango_hora),
                                                            fecha_marcacion=fecha,
                                                            hora_entrada=str(tiempo),
                                                            minutos_entrada_mas=minutos_ini_mas,
                                                            minutos_entrada_menos=minutos_ini_menos,
                                                            entrada_tardia=entrada_tardia
                                                                                        )   
                                                    msg="Entrada creada exitosamente de :",id_empleado," Rango de hora :",ho.rango_hora
                                        else:
                                            msg="Todavia no es hora de marcar",id_empleado," Rango de hora :",ho.rango_hora
                                        
                            else:
                                msg="Horario caducado de :",id_empleado
                        
                else:
                        msg="No tiene asignacion de horario el :",id_empleado
                                                            
            else:
                msg="Empleado esta inactivo"
        else:
            msg="Empleado con numero de Carnet",carnet," No existe"                
    return render (request, 'marcacion/menu_marcaciones/menu_marcaciones.html', {'msg':msg})

def entradatardia_o_salidatemprana (request):
    lista={}
    fecha1={}
    if request.method=='POST':
        fecha=request.POST['fecha']
        fecha1=datetime.strptime(fecha,"%Y-%m-%d")
        lista=Marcacion.objects.filter(fecha_marcacion=fecha)
        
    return render(request, 'marcacion/tardiasytemprana/lista_tardia_y_temprana.html',{'lista':lista,'fecha':fecha1})

def justificar(request,id):
    entrada=Marcacion.objects.get(id=id)
    return render(request,'marcacion/justificar/justificar.html',{'entrada':entrada})

def actualizarmarcacion(request):
    if request.method=='POST':
        etardia=request.POST['etardia']
        stemprana=request.POST['stemprana']
        id=request.POST['id']
        w = datetime.now()
        fecha = w.strftime('%Y-%m-%d')
        usu=request.POST['usu']
        just=request.POST['just']
        if len(just)>20:
            Marcacion.objects.filter(id=id).update(entrada_tardia=etardia,salida_temprana=stemprana,usuario_modifico=usu,fecha_modificacion=fecha,comentario=just)
            return redirect('menu_marcaciones')
        else:
            return redirect('justificar/'+id)

def listaausencias(request):
    if request.method=='POST':
        mes=request.POST['mes']
        mes1=datetime.strptime(mes,"%Y-%m")
        año=mes1.strftime('%Y')
        month=mes1.strftime('%m')
        lista=Ausencias.objects.filter(fecha__year=año).filter(fecha__month=month)
    return render(request,'marcacion/listaausencias/listaausencias.html',{'mes':mes1,'lista':lista})

def detalleausencia(request,id):
    ausencia=Ausencias.objects.get(id=id)
    return render(request,'marcacion/detalleausencia/detalleausencia.html',{'au':ausencia})

def actualizarausencia(request):
    if request.method=='POST':
        id=request.POST['id']
        falto=request.POST['esta']
        w = datetime.now()
        fecha = w.strftime('%Y-%m-%d')
        usu=request.POST['usu']
        just=request.POST['just']
        if len(just)>20:
            Ausencias.objects.filter(id=id).update(esta_faltando=falto,usuario_modfico=usu,fecha_modificacion=fecha,justificacion_ausencia=just)
            return redirect('menu_marcaciones')
        else:
            return redirect('detalleausencia/'+id)
    return redirect()

#seccion de reconocimiento facial 
def reg_camara_ip(request):
    if request.method=='POST':
        unidad_control=request.POST['unidad']
        url_camara_ip=request.POST['url']
        es_interna=request.POST['es']
        if es_interna=='si':
            Urlcamaraip.objects.create(unidad_control=unidad_control,url_camara_ip=url_camara_ip,es_interna=True)
            return redirect('menu_marcaciones')
        else:
            Urlcamaraip.objects.create(unidad_control=unidad_control,url_camara_ip=url_camara_ip)
            return redirect('menu_marcaciones')
    
    return render(request,'marcacion/reg_camara_ip/register_url.html')

@gzip.gzip_page
def recibir_url(request):
    #Esta funcion sirve para ejecutar el streaming de reconocimiento facial
    if request.method=='POST':
        id=request.POST['t']
        camara=Urlcamaraip.objects.get(id=id)
        url=camara.url_camara_ip
        try:
            return StreamingHttpResponse(gen2(Reconocimientofacial(url)),
                  content_type='multipart/x-mixed-replace; boundary=frame')
        except:
            pass
    return render(request,'inicio/index/monitor.html')

#-----------------------------------------------------------------------
def gen2(camreconcimiento):
    #Esta clase ejecuta la clase de reconocimiento facial exactamente la funcion de reconocimiento facial en monitor.py
    while True:
        frame = camreconcimiento.rec_facial()
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n\r\n')
        

#--------------------------------------------------------------------------


def reconocimientofacial(request):
    hora=""
    dataPath = os.getcwd()
    dataPath = os.path.join(dataPath, "CoreDatos")
    imagePaths = os.listdir(dataPath)
    face_recognizer = cv2.face.LBPHFaceRecognizer_create()
    face_recognizer.read('modeloLBPHFace.xml')
    faceClassif = cv2.CascadeClassifier(cv2.data.haarcascades+'haarcascade_frontalface_default.xml')
    video = cv2.VideoCapture(0)
    while(True):
        ret, frame = video.read()
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        auxFrame = gray.copy()
        faces = faceClassif.detectMultiScale(gray, 1.3, 5)
        for (x, y, w, h) in faces:
            rostro = auxFrame[y:y+h, x:x+w]
            rostro = cv2.resize(rostro, (150, 150),interpolation=cv2.INTER_CUBIC)
            result = face_recognizer.predict(rostro)
            if result[1] < 70:
                try:
                    now = datetime.now()
                    hora= now.strftime('%H:%M:%S')
                    print(result)
                    #con esto se resuelve que los valores esten fuera de rango
                    carnet =imagePaths[result[0]]
                    nombre=registrarentrada(carnet)
                    print(nombre)
                    #se hace consulta a la base de datos para obtener el nombre de la persona
                    #a quien pertenece el numero de carnet
                    cv2.putText(frame, '{}'.format(imagePaths[result[0]]),(x, y-25),2,1.1,(0, 255, 0),1,cv2.LINE_AA)
                    cv2.putText(frame, str(nombre)+" "+str(hora), (x+150, y-25), 2, 0.8, (255, 0, 255), 1, cv2.LINE_AA)
                    cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 255, 0), 4)
                    # obteniendo el nombre del reconocido para uso de BD
                    #print(result[1])
                                    
                except Exception as identifier:
                    print(identifier)
            else:
                cv2.putText(frame, 'Desconocido'+str(hora), (x, y-20), 2,0.8, (0, 0, 255), 1, cv2.LINE_AA)
                cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 0, 255), 2)

        cv2.imshow("Escriba y para finalizar", frame)
        if cv2.waitKey(1) & 0xFF == ord('y'):
            break
           
            
    video.release()
    cv2.destroyAllWindows()
    return  redirect('index')



def tardias_rango_fecha(request):
    if request.method=='POST':
        fecha_ini=request.POST['fecha1']
        fecha_fin=request.POST['fecha2']
        minutos=[]
        datos=[]
        if fecha_ini:
            if fecha_fin:
                fecha=datetime.strptime(fecha_fin,'%Y-%m-%d')
                nfecha=(fecha+timedelta(days=1))
                fechasumada=nfecha.strftime('%Y-%m-%d')
                empleado=Empleado.objects.all()
                marcacion=Marcacion.objects.exclude(fecha_marcacion__gte=fechasumada).filter(fecha_marcacion__gte=fecha_ini).order_by('fecha_marcacion')
                for em in empleado:
                    for marca in marcacion:
                        if marca.empleado_id==em.id:
                            if marca.entrada_tardia==True:
                                muni=marca.minutos_entrada_menos
                                m=datetime.strptime(muni,'%H:%M:%S')
                                minutos=timedelta(days=0,hours=m.hour, minutes=m.minute, seconds=m.second)
                                minutos+=minutos
                    
                    lista=(str(em),"tiempo sumado de llgadas tarde en rango de fecha desde: |",str(fecha_ini)," | hasta : | ",str(fecha_fin)," | Total minutos tarde : ",str(minutos))               
                    datos.append(lista)
                    
                return render( request,'marcacion/minutostarde/rangotardia.html',{'datos':datos})

def excelausencias(request):
    if request.method=='GET':
        fecha_ini=request.GET['fecha1']
        fecha_fin=request.GET['fecha2']
        if fecha_ini:
            if fecha_fin:
                fecha=datetime.strptime(fecha_fin,'%Y-%m-%d')
                nfecha=(fecha+timedelta(days=1))
                fechasumada=nfecha.strftime('%Y-%m-%d')
                marcacion=Marcacion.objects.exclude(fecha_marcacion__gte=fechasumada).filter(fecha_marcacion__gte=fecha_ini).order_by('fecha_marcacion')
                wb= Workbook()
                ws= wb.active
                ws['A1']='MARCACIONES POR RANGO DE FECHA, DESDE : '+str(fecha_ini)+' HASTA :'+str(fecha_fin)
                ws['A1'].alignment=Alignment(horizontal="center",vertical="center")
                ws['A1'].border= Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
                ws['A1'].fill=PatternFill(start_color='FFB833',end_color='FFB833',fill_type="solid")
                ws.merge_cells('A1:Q1')
                ws['A3']='CARNET'
                ws['B3']='NOMBRES'
                ws['C3']='APELLIDOS'
                ws['D3']='DEPARTAMENTO'
                ws['E3']='HORARIO'
                ws['F3']='RANGO DE HORA'
                ws['G3']='FECHA'
                ws['H3']='HORA ENTRADA'
                ws['I3']='MINUTOS MAS'
                ws['J3']='MINUTOS MENOS'
                ws['K3']='HORA SALIDA'
                ws['L3']='SALIDA MAS'
                ws['M3']='SALIDA MENOS'
                ws['N3']='FECHA MODIFICACION'
                ws['O3']='USUARIO MODIFICO'
                ws['P3']='COMENTARIO'
                ws['A3'].alignment=Alignment(horizontal="center",vertical="center")
                ws['A3'].border= Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
                ws['A3'].fill=PatternFill(start_color='33F0FF',end_color='33F0FF',fill_type="solid")
                ws['B3'].alignment=Alignment(horizontal="center",vertical="center")
                ws['B3'].border= Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
                ws['B3'].fill=PatternFill(start_color='33F0FF',end_color='33F0FF',fill_type="solid")
                ws['C3'].alignment=Alignment(horizontal="center",vertical="center")
                ws['C3'].border= Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
                ws['C3'].fill=PatternFill(start_color='33F0FF',end_color='33F0FF',fill_type="solid")
                ws['D3'].alignment=Alignment(horizontal="center",vertical="center")
                ws['D3'].border= Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
                ws['D3'].fill=PatternFill(start_color='33F0FF',end_color='33F0FF',fill_type="solid")
                ws['E3'].alignment=Alignment(horizontal="center",vertical="center")
                ws['E3'].border= Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
                ws['E3'].fill=PatternFill(start_color='33F0FF',end_color='33F0FF',fill_type="solid")
                ws['F3'].alignment=Alignment(horizontal="center",vertical="center")
                ws['F3'].border= Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
                ws['F3'].fill=PatternFill(start_color='33F0FF',end_color='33F0FF',fill_type="solid")
                ws['G3'].alignment=Alignment(horizontal="center",vertical="center")
                ws['G3'].border= Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
                ws['G3'].fill=PatternFill(start_color='33F0FF',end_color='33F0FF',fill_type="solid")
                ws['H3'].alignment=Alignment(horizontal="center",vertical="center")
                ws['H3'].border= Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
                ws['H3'].fill=PatternFill(start_color='33F0FF',end_color='33F0FF',fill_type="solid")
                ws['I3'].alignment=Alignment(horizontal="center",vertical="center")
                ws['I3'].border= Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
                ws['I3'].fill=PatternFill(start_color='33F0FF',end_color='33F0FF',fill_type="solid")
                ws['J3'].alignment=Alignment(horizontal="center",vertical="center")
                ws['J3'].border= Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
                ws['J3'].fill=PatternFill(start_color='33F0FF',end_color='33F0FF',fill_type="solid")
                ws['K3'].alignment=Alignment(horizontal="center",vertical="center")
                ws['K3'].border= Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
                ws['K3'].fill=PatternFill(start_color='33F0FF',end_color='33F0FF',fill_type="solid")
                ws['L3'].alignment=Alignment(horizontal="center",vertical="center")
                ws['L3'].border= Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
                ws['L3'].fill=PatternFill(start_color='33F0FF',end_color='33F0FF',fill_type="solid")
                ws['M3'].alignment=Alignment(horizontal="center",vertical="center")
                ws['M3'].border= Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
                ws['M3'].fill=PatternFill(start_color='33F0FF',end_color='33F0FF',fill_type="solid")
                ws['N3'].alignment=Alignment(horizontal="center",vertical="center")
                ws['N3'].border= Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
                ws['N3'].fill=PatternFill(start_color='33F0FF',end_color='33F0FF',fill_type="solid")
                ws['O3'].alignment=Alignment(horizontal="center",vertical="center")
                ws['O3'].border= Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
                ws['O3'].fill=PatternFill(start_color='33F0FF',end_color='33F0FF',fill_type="solid")
                ws['P3'].alignment=Alignment(horizontal="center",vertical="center")
                ws['P3'].border= Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
                ws['P3'].fill=PatternFill(start_color='33F0FF',end_color='33F0FF',fill_type="solid")
                cuenta=4
                for m in marcacion:
                    ws.cell(row=cuenta, column=1).value=m.empleado.carnet_empleado
                    ws.cell(row=cuenta, column=2).value=m.empleado.nombres
                    ws.cell(row=cuenta, column=3).value=m.empleado.apellidos
                    ws.cell(row=cuenta, column=4).value=m.empleado.departamento.nom_departamento
                    ws.cell(row=cuenta, column=5).value=m.horario.descripcion
                    ws.cell(row=cuenta, column=6).value=m.rango_horario
                    ws.cell(row=cuenta, column=7).value=m.fecha_marcacion
                    ws.cell(row=cuenta, column=8).value=m.hora_entrada
                    ws.cell(row=cuenta, column=9).value=m.minutos_entrada_mas
                    ws.cell(row=cuenta, column=10).value=m.minutos_entrada_menos
                    ws.cell(row=cuenta, column=11).value=m.hora_salida
                    ws.cell(row=cuenta, column=12).value=m.minutos_salida_mas
                    ws.cell(row=cuenta, column=13).value=m.minutos_salida_menos
                    ws.cell(row=cuenta, column=14).value=m.fecha_modificacion
                    ws.cell(row=cuenta, column=15).value=m.usuario_modifico
                    ws.cell(row=cuenta, column=16).value=m.comentario
                    
                    
                    cuenta+=1
                archivo="Marcaciones.xlsx"
                respuesta=HttpResponse(content_type="application/ms-excel")
                contenido="attachment; filename={0}".format(archivo)
                respuesta['Content-Disposition']=contenido
                wb.save(respuesta)
                return respuesta 