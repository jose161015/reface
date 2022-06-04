from evento.models import CodigoEvento, Evento
from django.shortcuts import redirect, render,HttpResponse
from empleado.models import Empleado
from datetime import date, datetime, timedelta
from openpyxl.styles.borders import Side
from openpyxl.styles.fills import PatternFill
from openpyxl.workbook.workbook import Workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment,Border

# Create your views here.
def menu_evento(request):
    return render (request, 'evento/menuevento/menuevento.html')


def regevento(request):
    lempleado=Empleado.objects.all()
    codigoevento=CodigoEvento.objects.all()
    if request.method=='POST':
        empleado=request.POST['empleado']
        cod_evento=request.POST['codevento']
        autorizado=request.POST['aupor']
        descripcion=request.POST['desc']
        fecha1=request.POST['fecha1']
        fecha2=request.POST['fecha2']
        usuario=request.POST['usu_reg']
        now = datetime.now()
        fereg=now.strftime('%Y-%m-%d')
        try:
            Evento.objects.create(
                empleado=Empleado.objects.get(id=empleado),
                codigo_evento=CodigoEvento.objects.get(id=cod_evento),
                justificacion=descripcion,
                autorizado_por=autorizado,
                fecha_inicio=fecha1,
                fecha_fin=fecha2,
                fecha_registro=fereg,
                usuario_registro=usuario
                )
            msgy="Evento guardado exitosamente"
            return render(request,'evento/regevento/regevento.html',{'lempleado':lempleado,'codigoevento':codigoevento,'msgy':msgy})
        except:
            msg="Evento no se guardo correctamente"
        return render(request,'evento/regevento/regevento.html',{'lempleado':lempleado,'codigoevento':codigoevento,'msg':msg})
    return render(request,'evento/regevento/regevento.html',{'lempleado':lempleado,'codigoevento':codigoevento})


def regcodevento(request):
    codevento=CodigoEvento.objects.all()
    if request.method=='POST':
        cod_evento=request.POST['codevento']
        desc=request.POST['desc']
        CodigoEvento.objects.create(cod_evento=cod_evento,descripcion_evento=desc)
    return render (request, 'evento/regcodevento/codevento.html',{'codevento':codevento})
def listarevento(request):
    eve=Evento.objects.all()
    return render(request, 'evento/listarevento/listarevento.html',{'evento':eve,'count':eve.count()})

def editarevento(request,id):
    event=Evento.objects.filter(id=id).first()
    return render(request,'evento/editarevento/editarevento.html',{'event':event,})


def actualizarevento(request):
    if request.method=='POST':
        id=request.POST['id']
        aupor=request.POST['aupor']
        desc=request.POST['desc']
        Evento.objects.filter(id=id).update(autorizado_por=aupor,justificacion=desc)
        return redirect('listarevento')

def eventorango(request):
    if request.method=='GET':
        f1=request.GET['fecha1']
        f2=request.GET['fecha2']  
        fecha=datetime.strptime(f2,'%Y-%m-%d')
        nfecha=(fecha+timedelta(days=1))
        fechasumada=nfecha.strftime('%Y-%m-%d')
        print(fechasumada)
        registro_rango=Evento.objects.exclude(fecha_inicio__gte=fechasumada).filter(fecha_fin__gte=f1).order_by('fecha_inicio')
        contenido={'registro_rango':registro_rango,'count':registro_rango.count(),'fecha1':f1,'fecha2':f2}
        return render(request,'eventorango.html',contenido)

def eventorangopdf(request):
    if request.method=='GET':
        f1=request.GET['fecha1']
        f2=request.GET['fecha2']
        fecha=datetime.strptime(f2,'%Y-%m-%d')
        nfecha=(fecha+timedelta(days=1))
        fechasumada=nfecha.strftime('%Y-%m-%d')
        registro_rango=Evento.objects.exclude(fecha_inicio__gte=fechasumada).filter(fecha_fin__gte=f1).order_by('fecha_inicio')
        wb= Workbook()
        ws= wb.active
        ws['A1']='EVENTOS POR RANGO DE FECHA, DESDE : '+str(f1)+' HASTA :'+str(f2)
        ws['A1'].alignment=Alignment(horizontal="center",vertical="center")
        ws['A1'].border= Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['A1'].fill=PatternFill(start_color='FFB833',end_color='FFB833',fill_type="solid")
        ws.merge_cells('A1:M1')
        hoy=datetime.now()
        ws['A2']='FECHA DE GENERACION DE DOCUMENTO : '+str(hoy)
        ws['A3']='CARNET'
        ws['B3']='NOMBRES'
        ws['C3']='APELLIDOS'
        ws['D3']='DEPARTAMENTO'
        ws['E3']='OCUPACION'
        ws['F3']='COD. EVENTO'
        ws['G3']='DESCRIPCION'
        ws['H3']='FECHA INICIO'
        ws['I3']='FECHA FINALIZACION'
        ws['J3']='RESPONSABLE QUE FIRMA'
        ws['K3']='MOTIVO'
        ws['L3']='FECHA REGISTRO'
        ws['M3']='RESPONSABLE DE REGISTRO'
        
        ws['A3'].alignment=Alignment(horizontal="center",vertical="center")
        ws['A3'].border= Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['A3'].fill=PatternFill(start_color='33F0FF',end_color='33F0FF',fill_type="solid")
        ws['B3'].alignment=Alignment(horizontal="center",vertical="center")
        ws['B3'].border= Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['B3'].fill=PatternFill(start_color='33F0FF',end_color='33F0FF',fill_type="solid")
        ws['C3'].alignment=Alignment(horizontal="center",vertical="center")
        ws['C3'].border= Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['C3'].fill=PatternFill(start_color='33F0FF',end_color='33F0FF',fill_type="solid")
        ws['D3'].alignment=Alignment(horizontal="center",vertical="center")
        ws['D3'].border= Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['D3'].fill=PatternFill(start_color='33F0FF',end_color='33F0FF',fill_type="solid")
        ws['E3'].alignment=Alignment(horizontal="center",vertical="center")
        ws['E3'].border= Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['E3'].fill=PatternFill(start_color='33F0FF',end_color='33F0FF',fill_type="solid")
        ws['F3'].alignment=Alignment(horizontal="center",vertical="center")
        ws['F3'].border= Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['F3'].fill=PatternFill(start_color='33F0FF',end_color='33F0FF',fill_type="solid")
        ws['G3'].alignment=Alignment(horizontal="center",vertical="center")
        ws['G3'].border= Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['G3'].fill=PatternFill(start_color='33F0FF',end_color='33F0FF',fill_type="solid")
        ws['H3'].alignment=Alignment(horizontal="center",vertical="center")
        ws['H3'].border= Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['H3'].fill=PatternFill(start_color='33F0FF',end_color='33F0FF',fill_type="solid")
        ws['I3'].alignment=Alignment(horizontal="center",vertical="center")
        ws['I3'].border= Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['I3'].fill=PatternFill(start_color='33F0FF',end_color='33F0FF',fill_type="solid")
        ws['J3'].alignment=Alignment(horizontal="center",vertical="center")
        ws['J3'].border= Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['J3'].fill=PatternFill(start_color='33F0FF',end_color='33F0FF',fill_type="solid")
        ws['K3'].alignment=Alignment(horizontal="center",vertical="center")
        ws['K3'].border= Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['K3'].fill=PatternFill(start_color='33F0FF',end_color='33F0FF',fill_type="solid")
        ws['L3'].alignment=Alignment(horizontal="center",vertical="center")
        ws['L3'].border= Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['L3'].fill=PatternFill(start_color='33F0FF',end_color='33F0FF',fill_type="solid")
        ws['M3'].alignment=Alignment(horizontal="center",vertical="center")
        ws['M3'].border= Border(left=Side(border_style="thin"),right=Side(border_style="thin"),top=Side(border_style="thin"),bottom=Side(border_style="thin"))
        ws['M3'].fill=PatternFill(start_color='33F0FF',end_color='33F0FF',fill_type="solid")
        cuenta=4
        for m in registro_rango:
            ws.cell(row=cuenta, column=1).value=m.empleado.carnet_empleado
            ws.cell(row=cuenta, column=2).value=m.empleado.nombres
            ws.cell(row=cuenta, column=3).value=m.empleado.apellidos
            ws.cell(row=cuenta, column=4).value=m.empleado.departamento.nom_departamento
            ws.cell(row=cuenta, column=5).value=m.empleado.ocupacion.nom_ocupacion
            ws.cell(row=cuenta, column=6).value=m.codigo_evento.cod_evento
            ws.cell(row=cuenta, column=7).value=m.codigo_evento.descripcion_evento
            ws.cell(row=cuenta, column=8).value=m.fecha_inicio
            ws.cell(row=cuenta, column=9).value=m.fecha_fin
            ws.cell(row=cuenta, column=10).value=m.autorizado_por
            ws.cell(row=cuenta, column=11).value=m.justificacion
            ws.cell(row=cuenta, column=12).value=m.fecha_registro
            ws.cell(row=cuenta, column=13).value=m.usuario_registro       
            cuenta+=1
        archivo="Eventos-"+str(f1)+"-"+str(f2)+".xlsx"
        respuesta=HttpResponse(content_type="application/ms-excel")
        contenido="attachment; filename={0}".format(archivo)
        respuesta['Content-Disposition']=contenido
        wb.save(respuesta)
        return respuesta 
        